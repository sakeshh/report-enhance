from __future__ import annotations

import csv
import json
import os
import re
import tempfile
import uuid
from dataclasses import dataclass, field
from typing import Any, Callable, List, Optional, Tuple

_VIEW_SAFE = re.compile(r"[^a-zA-Z0-9_]+")


class SourceConnectionError(RuntimeError):
    """Raised when a source cannot be connected or a DuckDB VIEW cannot be registered."""


@dataclass
class SourceConnection:
    """
    Handle for a single assessed source.

    For file-backed DuckDB sources, ``duckdb_database_path`` is set so GX can open the
    same database via SQLAlchemy after ``prepare_duckdb_for_gx_validation`` closes ``con``.
    """

    con: Any  # duckdb.DuckDBPyConnection when using DuckDB; None for azure_sql
    view_name: str
    source_type: str
    source_path: str
    dataset_name: str
    table_name: Optional[str] = None
    duckdb_database_path: Optional[str] = None
    _tmp_paths: List[str] = field(default_factory=list)
    _on_close: List[Callable[[], None]] = field(default_factory=list)

    def __post_init__(self) -> None:
        if self._on_close is None:
            self._on_close = []


def _log(msg: str) -> None:
    print(f"[source_connector] {msg}")


def _safe_view_base(dataset_name: str) -> str:
    base = _VIEW_SAFE.sub("_", dataset_name or "dataset").strip("_") or "dataset"
    return base[:48]


def _new_view_name(dataset_name: str) -> str:
    return f"v_{uuid.uuid4().hex[:10]}_{_safe_view_base(dataset_name)}"


def _parse_az_path(az_path: str) -> Tuple[str, str]:
    if not az_path.startswith("az://"):
        raise SourceConnectionError(f"Expected az:// path, got: {az_path!r}")
    rest = az_path[5:]
    if "/" not in rest:
        raise SourceConnectionError(f"Invalid az:// path (missing blob key): {az_path!r}")
    container, blob_key = rest.split("/", 1)
    if not container or not blob_key:
        raise SourceConnectionError(f"Invalid az:// path: {az_path!r}")
    return container, blob_key


def _azure_connector_cfg(source_config: dict, container: str) -> dict:
    account = (
        source_config.get("azure_account_name")
        or os.environ.get("AZURE_STORAGE_ACCOUNT_NAME")
        or ""
    ).strip()
    key = (
        source_config.get("azure_account_key")
        or os.environ.get("AZURE_STORAGE_ACCOUNT_KEY")
        or ""
    ).strip()
    if not account or not key:
        raise SourceConnectionError(
            "Azure storage credentials missing: set azure_account_name / azure_account_key "
            "on the source config or AZURE_STORAGE_ACCOUNT_NAME / AZURE_STORAGE_ACCOUNT_KEY in the environment."
        )
    return {"account_name": account, "account_key": key, "container": container}


def _parquet_https_url(
    az_path: str,
    *,
    account_name: str,
    account_key: str,
) -> str:
    from datetime import datetime, timedelta, timezone

    try:
        from azure.storage.blob import BlobSasPermissions, generate_blob_sas  # type: ignore[import]
    except ImportError as e:
        raise SourceConnectionError(
            "Parquet on Azure requires azure-storage-blob. Install with: pip install azure-storage-blob"
        ) from e

    container, blob_key = _parse_az_path(az_path)
    expiry = datetime.now(timezone.utc) + timedelta(hours=2)
    sas = generate_blob_sas(
        account_name=account_name,
        container_name=container,
        blob_name=blob_key,
        account_key=account_key,
        permission=BlobSasPermissions(read=True),
        expiry=expiry,
    )
    from urllib.parse import quote

    encoded_blob = "/".join(quote(seg, safe="") for seg in blob_key.split("/"))
    return f"https://{account_name}.blob.core.windows.net/{container}/{encoded_blob}?{sas}"


def _sql_escape_literal(s: str) -> str:
    return s.replace("'", "''")


def prepare_duckdb_for_gx_validation(conn: SourceConnection) -> None:
    """
    Close the native DuckDB connection so SQLAlchemy can open the same on-disk database.

    Safe to call for non-DuckDB sources (no-op).
    """
    if conn.duckdb_database_path and conn.con is not None:
        try:
            conn.con.close()
        finally:
            conn.con = None
        _log(f"released native DuckDB lock for GX (dataset={conn.dataset_name})")


def connect_source(source_config: dict) -> SourceConnection:
    """
    Build a per-job DuckDB VIEW (or Azure SQL handle) for downstream SQL and GX.

    ``source_config`` keys match the Agent Dhara contract (type, path, dataset_name, ...).
    """
    cfg = source_config or {}
    source_type = str(cfg.get("type") or "").lower().strip()
    path = str(cfg.get("path") or "").strip()
    dataset_name = str(cfg.get("dataset_name") or "").strip()
    if not source_type:
        raise SourceConnectionError("source_config.type is required")
    if not path:
        raise SourceConnectionError("source_config.path is required")
    if not dataset_name:
        raise SourceConnectionError("source_config.dataset_name is required")

    view_name = _new_view_name(dataset_name)
    tmp_paths: List[str] = []
    on_close: List[Callable[[], None]] = []

    if source_type == "azure_sql":
        table_name = (cfg.get("table_name") or "").strip()
        if not table_name:
            raise SourceConnectionError("azure_sql source requires table_name")
        _log(f"registered azure_sql handle for {path!r} table={table_name}")
        return SourceConnection(
            con=None,
            view_name=table_name,
            source_type="azure_sql",
            source_path=path,
            dataset_name=dataset_name,
            table_name=table_name,
            duckdb_database_path=None,
            _tmp_paths=tmp_paths,
            _on_close=on_close,
        )

    import duckdb  # lazy import

    db_path = os.path.join(tempfile.gettempdir(), f"dhara_duck_{uuid.uuid4().hex}.duckdb")
    con = duckdb.connect(db_path)

    def _cleanup_db_file() -> None:
        try:
            if os.path.isfile(db_path):
                os.unlink(db_path)
        except OSError:
            pass

    on_close.append(_cleanup_db_file)

    try:
        if source_type == "csv":
            from connectors.azure_blob_storage import AzureBlobStorageConnector

            if not path.startswith("az://"):
                raise SourceConnectionError("CSV path must be az://container/blob for Azure download")
            container, blob_key = _parse_az_path(path)
            connector = AzureBlobStorageConnector(_azure_connector_cfg(cfg, container))
            tmp_csv = os.path.join(tempfile.gettempdir(), f"dhara_{uuid.uuid4().hex}.csv")
            try:
                connector.download_blob_to_file(blob_key, tmp_csv)
            except Exception as e:
                raise SourceConnectionError(f"Blob download failed for {path}: {e}") from e
            tmp_paths.append(tmp_csv)
            on_close.append(lambda p=tmp_csv: _unlink_quiet(p))

            p_esc = _sql_escape_literal(tmp_csv.replace("\\", "/"))
            con.execute(
                f'CREATE VIEW "{view_name}" AS SELECT * FROM read_csv_auto(\'{p_esc}\', header=true)'
            )

        elif source_type == "json":
            from connectors.azure_blob_storage import AzureBlobStorageConnector

            if not path.startswith("az://"):
                raise SourceConnectionError("JSON path must be az://container/blob for Azure download")
            container, blob_key = _parse_az_path(path)
            connector = AzureBlobStorageConnector(_azure_connector_cfg(cfg, container))
            tmp_json = os.path.join(tempfile.gettempdir(), f"dhara_{uuid.uuid4().hex}.json")
            try:
                connector.download_blob_to_file(blob_key, tmp_json)
            except Exception as e:
                raise SourceConnectionError(f"Blob download failed for {path}: {e}") from e
            tmp_paths.append(tmp_json)
            on_close.append(lambda p=tmp_json: _unlink_quiet(p))

            p_esc = _sql_escape_literal(tmp_json.replace("\\", "/"))
            con.execute(f'CREATE VIEW "{view_name}" AS SELECT * FROM read_json_auto(\'{p_esc}\')')

        elif source_type == "parquet":
            try:
                import adlfs  # noqa: F401
                import fsspec  # noqa: F401
            except ImportError:
                _log("optional adlfs/fsspec import failed (SAS parquet path does not require them)")
            account = (
                cfg.get("azure_account_name") or os.environ.get("AZURE_STORAGE_ACCOUNT_NAME") or ""
            ).strip()
            key = (cfg.get("azure_account_key") or os.environ.get("AZURE_STORAGE_ACCOUNT_KEY") or "").strip()
            if not account or not key:
                raise SourceConnectionError("Parquet az:// paths require azure account name/key")
            https_url = _parquet_https_url(path, account_name=account, account_key=key)
            u_esc = _sql_escape_literal(https_url)
            con.execute(f'CREATE VIEW "{view_name}" AS SELECT * FROM read_parquet(\'{u_esc}\')')

        elif source_type == "xml":
            try:
                import xmltodict  # type: ignore[import]
            except ImportError as e:
                raise SourceConnectionError("XML sources require xmltodict. pip install xmltodict") from e
            from connectors.azure_blob_storage import AzureBlobStorageConnector

            if not path.startswith("az://"):
                raise SourceConnectionError("XML path must be az://container/blob")
            container, blob_key = _parse_az_path(path)
            connector = AzureBlobStorageConnector(_azure_connector_cfg(cfg, container))
            try:
                raw = connector._download_blob_bytes(blob_key)  # noqa: SLF001 — internal reuse
            except Exception as e:
                raise SourceConnectionError(f"Blob download failed for {path}: {e}") from e
            try:
                parsed = xmltodict.parse(raw.decode("utf-8", errors="replace"))
            except Exception as e:
                raise SourceConnectionError(f"XML parse failed for {path}: {e}") from e

            root_key = cfg.get("root_key")
            if root_key and isinstance(parsed, dict) and root_key in parsed:
                parsed = parsed[root_key]

            if isinstance(parsed, dict) and len(parsed) == 1:
                sole_key = next(iter(parsed))
                sole_val = parsed[sole_key]
                if isinstance(sole_val, list):
                    rows = sole_val
                elif isinstance(sole_val, dict):
                    rows = [sole_val]
                else:
                    rows = [{"value": sole_val}]
            elif isinstance(parsed, list):
                rows = parsed
            else:
                rows = [parsed] if parsed is not None else []

            tmp_json = os.path.join(tempfile.gettempdir(), f"dhara_{uuid.uuid4().hex}.json")
            with open(tmp_json, "w", encoding="utf-8") as f:
                json.dump(rows, f)
            tmp_paths.append(tmp_json)
            on_close.append(lambda p=tmp_json: _unlink_quiet(p))

            p_esc = _sql_escape_literal(tmp_json.replace("\\", "/"))
            con.execute(f'CREATE VIEW "{view_name}" AS SELECT * FROM read_json_auto(\'{p_esc}\')')

        elif source_type == "excel":
            try:
                from openpyxl import load_workbook  # type: ignore[import]
            except ImportError as e:
                raise SourceConnectionError("Excel sources require openpyxl. pip install openpyxl") from e
            from connectors.azure_blob_storage import AzureBlobStorageConnector

            if not path.startswith("az://"):
                raise SourceConnectionError("Excel path must be az://container/blob")
            container, blob_key = _parse_az_path(path)
            connector = AzureBlobStorageConnector(_azure_connector_cfg(cfg, container))
            tmp_xlsx = os.path.join(tempfile.gettempdir(), f"dhara_{uuid.uuid4().hex}.xlsx")
            try:
                connector.download_blob_to_file(blob_key, tmp_xlsx)
            except Exception as e:
                raise SourceConnectionError(f"Blob download failed for {path}: {e}") from e
            tmp_paths.append(tmp_xlsx)
            on_close.append(lambda p=tmp_xlsx: _unlink_quiet(p))

            tmp_csv = os.path.join(tempfile.gettempdir(), f"dhara_{uuid.uuid4().hex}.csv")
            wb = load_workbook(tmp_xlsx, read_only=True, data_only=True)
            try:
                sheet_name = cfg.get("sheet_name")
                if sheet_name:
                    ws = wb[str(sheet_name)]
                else:
                    ws = wb.active
                with open(tmp_csv, "w", encoding="utf-8", newline="") as fcsv:
                    w = csv.writer(fcsv)
                    for row in ws.iter_rows(values_only=True):
                        w.writerow(["" if c is None else c for c in row])
            finally:
                wb.close()
            tmp_paths.append(tmp_csv)
            on_close.append(lambda p=tmp_csv: _unlink_quiet(p))

            p_esc = _sql_escape_literal(tmp_csv.replace("\\", "/"))
            con.execute(
                f'CREATE VIEW "{view_name}" AS SELECT * FROM read_csv_auto(\'{p_esc}\', header=true)'
            )

        else:
            con.close()
            for fn in reversed(on_close):
                try:
                    fn()
                except Exception:
                    pass
            raise SourceConnectionError(f"Unsupported source type: {source_type!r}")

    except SourceConnectionError:
        con.close()
        for fn in reversed(on_close):
            try:
                fn()
            except Exception:
                pass
        raise
    except Exception as e:
        con.close()
        for fn in reversed(on_close):
            try:
                fn()
            except Exception:
                pass
        raise SourceConnectionError(f"DuckDB VIEW creation failed: {e}") from e

    _log(f"registered VIEW {view_name} for {path}")
    return SourceConnection(
        con=con,
        view_name=view_name,
        source_type=source_type,
        source_path=path,
        dataset_name=dataset_name,
        table_name=None,
        duckdb_database_path=db_path,
        _tmp_paths=tmp_paths,
        _on_close=on_close,
    )


def _unlink_quiet(p: str) -> None:
    try:
        if os.path.isfile(p):
            os.unlink(p)
    except OSError:
        pass


def validate_view_accessible(conn: SourceConnection) -> bool:
    """Run ``SELECT COUNT(*) FROM {view} LIMIT 1`` (or SQL Server equivalent for azure_sql)."""
    if conn.source_type == "azure_sql":
        if not conn.table_name:
            return False
        try:
            from sqlalchemy import create_engine, text  # lazy import
        except ImportError as e:
            raise SourceConnectionError("azure_sql validation requires sqlalchemy. pip install sqlalchemy") from e

        eng = create_engine(conn.source_path)
        tbl = conn.table_name
        with eng.connect() as c:
            c.execute(text(f"SELECT COUNT(*) FROM {tbl} WHERE 1=1"))
        return True

    if conn.con is None:
        return False
    try:
        conn.con.execute(f'SELECT COUNT(*) FROM "{conn.view_name}" LIMIT 1')
        return True
    except Exception as e:
        raise SourceConnectionError(f"VIEW {conn.view_name!r} is not accessible: {e}") from e


def close_source(conn: SourceConnection) -> None:
    """Close DuckDB (if open), dispose GX engines are caller responsibility; delete temp artifacts."""
    if conn.con is not None:
        try:
            conn.con.close()
        except Exception:
            pass
        conn.con = None
    for fn in reversed(conn._on_close):
        try:
            fn()
        except Exception:
            pass
    conn._on_close.clear()
    for p in conn._tmp_paths:
        _unlink_quiet(p)
    conn._tmp_paths.clear()
    _log(f"closed source dataset={conn.dataset_name!r}")
